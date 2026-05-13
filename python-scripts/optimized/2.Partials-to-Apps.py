import pandas as pd
from datetime import datetime
import os
import warnings
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    module="openpyxl"
)

# ==================================================
# CONFIG
# ==================================================

today_str = datetime.today().strftime("%m-%d-%y")

base_path = r"C:\Users\Emanuel\PyCharmMiscProject\WatsonInstitute"

reports_path = os.path.join(
    base_path,
    "reports"
)

salesforce_path = os.path.join(
    base_path,
    "salesforce"
)

# --------------------------------------------------
# OUTPUT FOLDER
# --------------------------------------------------

output_path = os.path.join(
    base_path,
    "Partial-Entries-Converted-to-Apps"
)

os.makedirs(output_path, exist_ok=True)

# ==================================================
# HELPERS
# ==================================================

def normalize(text):

    return re.sub(
        r"[^a-z0-9]",
        "",
        str(text).lower()
    )


# --------------------------------------------------
# EXTRACT DATE
# --------------------------------------------------

def extract_date(file_name):

    match = re.match(
        r"(\d{2}-\d{2}-\d{4})",
        file_name
    )

    if not match:
        return None

    try:

        return datetime.strptime(
            match.group(1),
            "%m-%d-%Y"
        )

    except:
        return None


# --------------------------------------------------
# FIND MATCHING SALESFORCE FILE
# --------------------------------------------------

def find_matching_salesforce_file(program_name):

    normalized_program = normalize(program_name)

    for sf_file in os.listdir(salesforce_path):

        if not sf_file.endswith(".xlsx"):
            continue

        sf_norm = normalize(sf_file)

        if (
            normalized_program in sf_norm
            or sf_norm in normalized_program
        ):

            return os.path.join(
                salesforce_path,
                sf_file
            )

    return None


# --------------------------------------------------
# AUTO FORMAT EXCEL
# --------------------------------------------------

def auto_adjust_columns(file_path):

    wb = load_workbook(file_path)

    ws = wb.active

    # Auto column width
    for column_cells in ws.columns:

        max_length = 0

        column = column_cells[0].column

        column_letter = get_column_letter(column)

        for cell in column_cells:

            try:

                if cell.value is not None:

                    cell_value = str(cell.value)

                    max_length = max(
                        max_length,
                        len(cell_value)
                    )

                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="top"
                    )

            except:
                pass

        adjusted_width = min(
            (max_length + 4) * 1.1,
            60
        )

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width

    # Row heights
    for row in ws.iter_rows():

        max_lines = 1

        for cell in row:

            if cell.value:

                lines = (
                    str(cell.value).count("\n") + 1
                )

                max_lines = max(
                    max_lines,
                    lines
                )

        ws.row_dimensions[
            row[0].row
        ].height = max(
            18,
            max_lines * 14
        )

    wb.save(file_path)


# --------------------------------------------------
# SAFE COLUMN
# --------------------------------------------------

def safe_column(df, col_name):

    if col_name not in df.columns:

        print(f"Missing column: {col_name}")

        df[col_name] = None

    return df[col_name]


# ==================================================
# FIND REPORTS BY PROGRAM
# ==================================================

program_reports = {}

# --------------------------------------------------
# IMPORTANT:
# ONLY ORIGINAL REPORTS
# EXCLUDE WEEKLY FILES
# --------------------------------------------------

pattern = re.compile(
    r"(\d{2}-\d{2}-\d{4}) - (.+?) - Partial Entries Report\.xlsx$",
    re.IGNORECASE
)

# --------------------------------------------------
# SEARCH RECURSIVELY
# --------------------------------------------------

for root, dirs, files in os.walk(reports_path):

    for file in files:

        if not file.endswith(".xlsx"):
            continue

        # EXCLUDE WEEKLY FILES
        if " - Weekly.xlsx" in file:
            continue

        match = pattern.match(file)

        if not match:
            continue

        date_str, program_name = match.groups()

        file_date = extract_date(file)

        if not file_date:
            continue

        normalized_program = normalize(
            program_name
        )

        full_path = os.path.join(
            root,
            file
        )

        if normalized_program not in program_reports:

            program_reports[
                normalized_program
            ] = {
                "display_name": program_name,
                "files": []
            }

        program_reports[
            normalized_program
        ]["files"].append(
            (file_date, full_path, file)
        )

# ==================================================
# MAIN PROCESS
# ==================================================

total_converted = 0

summary = []

for program_key, data in program_reports.items():

    display_name = data["display_name"]

    files = sorted(
        data["files"],
        key=lambda x: x[0],
        reverse=True
    )

    # Need at least 2 reports
    if len(files) < 2:
        continue

    # --------------------------------------------------
    # LATEST REPORT
    # --------------------------------------------------

    latest_date, latest_path, latest_file = files[0]

    # --------------------------------------------------
    # PREVIOUS REPORT
    # --------------------------------------------------

    previous_date, previous_path, previous_file = files[1]

    print("\n========================================")
    print(f"PROGRAM:  {display_name}")
    print(f"LATEST:   {latest_file}")
    print(f"PREVIOUS: {previous_file}")
    print("========================================")

    # --------------------------------------------------
    # FIND SALESFORCE FILE
    # --------------------------------------------------

    salesforce_file = find_matching_salesforce_file(
        display_name
    )

    if not salesforce_file:

        print("No Salesforce match found.")
        continue

    print(f"SALESFORCE: {os.path.basename(salesforce_file)}")

    # --------------------------------------------------
    # READ FILES
    # --------------------------------------------------

    try:

        df_latest = pd.read_excel(
            latest_path,
            sheet_name="Final"
        )

        df_previous = pd.read_excel(
            previous_path,
            sheet_name="Final"
        )

        df_salesforce = pd.read_excel(
            salesforce_file
        )

    except Exception as e:

        print(f"Error reading files: {e}")

        continue

    # ==================================================
    # CLEAN EMAILS
    # ==================================================

    latest_email_col = "Email (Enter Email)"

    previous_email_col = "Email (Enter Email)"

    salesforce_email_col = "Email"

    # Current partials
    df_latest["Email_clean"] = (
        safe_column(
            df_latest,
            latest_email_col
        )
        .astype(str)
        .str.strip()
        .str.lower()
    )

    # Previous partials
    df_previous["Email_clean"] = (
        safe_column(
            df_previous,
            previous_email_col
        )
        .astype(str)
        .str.strip()
        .str.lower()
    )

    # Salesforce emails
    df_salesforce["Email_clean"] = (
        safe_column(
            df_salesforce,
            salesforce_email_col
        )
        .astype(str)
        .str.strip()
        .str.lower()
    )

    # Remove invalid emails
    invalid_values = {
        "",
        "nan",
        "none",
        "null"
    }

    df_latest = df_latest[
        ~df_latest["Email_clean"].isin(
            invalid_values
        )
    ]

    df_previous = df_previous[
        ~df_previous["Email_clean"].isin(
            invalid_values
        )
    ]

    df_salesforce = df_salesforce[
        ~df_salesforce["Email_clean"].isin(
            invalid_values
        )
    ]

    # ==================================================
    # FIND PEOPLE WHO LEFT PARTIALS
    # ==================================================

    previous_emails = set(
        df_previous["Email_clean"]
    )

    latest_emails = set(
        df_latest["Email_clean"]
    )

    # People who disappeared
    removed_emails = previous_emails - latest_emails

    print(f"Previous partials: {len(previous_emails)}")
    print(f"Current partials:  {len(latest_emails)}")
    print(f"Removed partials:  {len(removed_emails)}")

    # --------------------------------------------------
    # FILTER REMOVED ROWS
    # --------------------------------------------------

    df_removed = df_previous[
        df_previous["Email_clean"].isin(
            removed_emails
        )
    ]

    # ==================================================
    # FIND CONVERSIONS
    # ==================================================
    #
    # Logic:
    #
    # 1. Was partial last week
    # 2. Is no longer partial this week
    # 3. Exists in Salesforce
    #
    # ==================================================

    merged = pd.merge(
        df_removed,
        df_salesforce,
        on="Email_clean",
        how="inner"
    )

    # Remove duplicate conversions
    merged = merged.drop_duplicates(
        subset=["Email_clean"]
    )

    converted_count = len(merged)

    total_converted += converted_count

    print(
        f"Partial Entries Converted to Apps This Week: {converted_count}"
    )

    if converted_count == 0:
        continue

    # ==================================================
    # FINAL OUTPUT
    # ==================================================

    final_df = pd.DataFrame({

        "Contact ID": safe_column(
            merged,
            "Contact ID"
        ),

        "First Name": safe_column(
            merged,
            "First Name"
        ),

        "Last Name": safe_column(
            merged,
            "Last Name"
        ),

        "Email": safe_column(
            merged,
            "Email"
        ),

        "Application Date Submitted": safe_column(
            merged,
            "Application Date Submitted"
        ),

    })

    # --------------------------------------------------
    # FORMAT DATES
    # --------------------------------------------------

    final_df["Application Date Submitted"] = (
        pd.to_datetime(
            final_df["Application Date Submitted"],
            errors="coerce"
        )
    )

    final_df = final_df.sort_values(
        by="Application Date Submitted"
    )

    final_df["Application Date Submitted"] = (
        final_df["Application Date Submitted"]
        .dt.strftime("%m-%d-%Y")
    )

    # ==================================================
    # OUTPUT FILE NAME
    # ==================================================

    clean_program_name = normalize(
        display_name
    )

    output_file = (
        f"{today_str}-"
        f"{clean_program_name}-"
        f"partials-converted-to-apps-this-week.xlsx"
    )

    output_full_path = os.path.join(
        output_path,
        output_file
    )

    # Remove existing file if open/exists
    if os.path.exists(output_full_path):

        try:
            os.remove(output_full_path)

        except PermissionError:

            print(
                f"Close Excel file first: {output_file}"
            )

            continue

    # ==================================================
    # EXPORT
    # ==================================================

    try:

        final_df.to_excel(
            output_full_path,
            index=False
        )

        auto_adjust_columns(
            output_full_path
        )

        print(f"Created: {output_file}")

        summary.append((
            display_name,
            converted_count,
            output_file
        ))

    except Exception as e:

        print(f"Error writing file: {e}")

# ==================================================
# SUMMARY
# ==================================================

print("\n========================================")
print("SUMMARY")
print("========================================\n")

for program, count, file in summary:

    print(program)

    print(
        f"Partial Entries Converted to Apps This Week: {count}"
    )

    print(f"Created: {file}\n")



print("\nAll files processed.")
